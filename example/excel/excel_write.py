import os
from openpyxl import Workbook

# 폴더가 없으면 폴더생성
if not os.path.exists("data"):
    os.mkdir("data")

# "Example" 이름으로 시트를 생성
wb = Workbook()
ws = wb.create_sheet("Example")

# 헤더 입력
ws['A1'] = 'Code'
ws['B1'] = 'Name'
ws['C1'] = 'Memo'
item4 = "아이템4"

# 행 단위로 추가
ws.append(['A1', 'item1', 20.2])
ws.append(['B2', 'item2', 'test'])
ws.append(['C3', 'item3', '실험'])
ws.append(['D4', item4, '2020-08-24'])

# 셀 단위로 추가
# ws.cell(row, column, value)
ws.cell(7, 1, '7행 1열')
ws.cell(7, 3, '7행 3열')
ws.cell(7, 5, '7행 5열')

# Excel 파일로 저장하기
wb.save('data/sample.xlsx')
wb.close()
