import os
from openpyxl import load_workbook

file_path = "data/sample.xlsx"

# 엑셀파일을 읽는 load_workbook 함수
# data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
wb = load_workbook(file_path, data_only=True)

# 시트 이름으로 불러오기
ws = wb['Example']

# 엑셀의 셀주소로 값을 출력
print(ws['A1'].value)
print(ws['B2'].value)
print(ws['C3'].value)
print()

# 셀 좌표로 값 출력
print(ws.cell(2, 1).value)
print(ws.cell(2, 2).value)
print(ws.cell(2, 3).value)
print()

records = []
for row in ws.rows:
    if row[0].value is not None:
        records.append((row[0].value, row[1].value, row[2].value))
    else:
        continue

for record in records:
    print(record)

wb.close()
